import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter
import logging
from logging import Logger
from pathlib import Path
from datetime import datetime
from typing import Optional, Tuple
import warnings

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QTextEdit, QFileDialog,
    QMessageBox, QGroupBox, QGridLayout, QProgressBar
)
from PyQt6.QtCore import QThread, pyqtSignal
from PyQt6.QtGui import QFont


# 抑制pandas警告
warnings.filterwarnings('ignore', category=pd.errors.PerformanceWarning)

class ConfigManager:
    """配置管理类"""
    
    # 默认配置
    DEFAULT_CONFIG = {
        'target_types': ['[21]当年预算', '[22]上年结转（非权责制）', '[23]上年结余（非权责制）'],
        'fund_nature_code': 1,  # 政府预算资金编码第一位为1
        'excluded_unit_code': 9,  # 排除的预算单位编码第一位为9
        'high_deviation_threshold': 0.1,  # 高偏离度阈值10%
        'decimal_places': 6,  # 差额保留小数位数
        'percentage_format': '0.00%',  # 百分比格式
        'max_column_width': 30,  # 最大列宽
        'column_width_padding': 2,  # 列宽填充
    }
    
    @classmethod
    def get_config(cls, key: str, default=None):
        """获取配置值"""
        return cls.DEFAULT_CONFIG.get(key, default)
    
    @classmethod
    def set_config(cls, key: str, value):
        """设置配置值"""
        if key in cls.DEFAULT_CONFIG:
            cls.DEFAULT_CONFIG[key] = value

from common.logger import get_logger, add_qt_signal

class DataProcessor:
    """数据处理基类"""
    
    def __init__(self, logger: Logger):
        self.logger = logger
    
    @staticmethod
    def safe_numeric_conversion(series: pd.Series, errors: str = 'coerce') -> pd.Series:
        """安全的数值转换"""
        return pd.to_numeric(series, errors=errors)
    
    @staticmethod
    def clean_numeric_string(series: pd.Series) -> pd.Series:
        """清理数值字符串（去除逗号等）"""
        return series.astype(str).str.replace(',', '').str.replace(' ', '')
    
    def validate_dataframe(self, df: pd.DataFrame, required_columns: list) -> bool:
        """验证DataFrame是否包含必需的列"""
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            self.logger.error(f"缺少必需的列: {missing_columns}")
            return False
        return True

class BudgetExecutionProcessor(DataProcessor):
    """预算执行数据处理器"""
    
    REQUIRED_COLUMNS = ['预算单位', '指标类型', '资金性质', '集中支付_实际支出数(非政采)', 
                       '集中支付_实际支出数（政采）', '集中支付_转列支出(非政采)', 
                       '集中支付_转列支出（政采）', '实拨_实际支出']
    
    def process(self, df: pd.DataFrame) -> pd.DataFrame:
        """处理预算执行数据"""
        if not self.validate_dataframe(df, self.REQUIRED_COLUMNS):
            raise ValueError("预算执行数据格式不正确")
        
        self.logger.info("开始处理预算执行数据...")
        original_count = len(df)
        
        # 数据清洗和转换
        df = self._clean_and_transform(df)
        
        # 数据筛选
        df = self._filter_data(df)
        
        # 计算支出数
        df = self._calculate_expenditure(df)
        
        # 分组汇总
        df = self._group_and_aggregate(df)
        
        # 添加辅助列
        df = self._add_helper_columns(df)
        
        filtered_count = len(df)
        self.logger.info(f"预算执行数据处理完成: {original_count} -> {filtered_count} 条记录")
        
        return df
    
    def _clean_and_transform(self, df: pd.DataFrame) -> pd.DataFrame:
        """清洗和转换数据"""
        # 提取资金性质编码
        df['资金性质编码'] = self.safe_numeric_conversion(
            df['资金性质'].str.slice(1, 2)
        )
        
        # 提取预算单位编码
        df['预算单位编码'] = self.safe_numeric_conversion(
            df['预算单位'].str.slice(1, 2)
        ).astype('Int64')
        
        return df
    
    def _filter_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """筛选数据"""
        target_types = ConfigManager.get_config('target_types')
        fund_nature_code = ConfigManager.get_config('fund_nature_code')
        excluded_unit_code = ConfigManager.get_config('excluded_unit_code')
        
        return df[
            (df['预算单位'] != '0') &
            (df['预算单位'] != 0) &
            (df['指标类型'].isin(target_types)) &
            (df['资金性质编码'] == fund_nature_code) &
            (df['预算单位编码'] != excluded_unit_code)
        ].copy()
    
    def _calculate_expenditure(self, df: pd.DataFrame) -> pd.DataFrame:
        """计算支出数"""
        expenditure_columns = [
            '集中支付_实际支出数(非政采)', '集中支付_实际支出数（政采）',
            '集中支付_转列支出(非政采)', '集中支付_转列支出（政采）', '实拨_实际支出'
        ]
        
        # 确保所有列都是数值类型
        for col in expenditure_columns:
            df[col] = self.safe_numeric_conversion(df[col]).fillna(0)
        
        df['预算执行_支出数'] = df[expenditure_columns].sum(axis=1)
        return df
    
    def _group_and_aggregate(self, df: pd.DataFrame) -> pd.DataFrame:
        """分组并汇总"""
        return df.groupby('预算单位', as_index=False)['预算执行_支出数'].sum()
    
    def _add_helper_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """添加辅助列"""
        df['单位编码'] = self.safe_numeric_conversion(
            df['预算单位'].str.slice(1, 7)
        )
        df['序号'] = range(1, len(df) + 1)
        return df

class AccountingProcessor(DataProcessor):
    """会计核算数据处理器"""
    
    REQUIRED_COLUMNS = ['账套', '借方累计', '贷方累计']
    
    def process(self, df: pd.DataFrame) -> pd.DataFrame:
        """处理会计核算数据"""
        if not self.validate_dataframe(df, self.REQUIRED_COLUMNS):
            raise ValueError("会计核算数据格式不正确")
        
        self.logger.info("开始处理会计核算数据...")
        original_count = len(df)
        
        # 数据清洗
        df = self._clean_data(df)
        
        # 数值转换
        df = self._convert_numeric_columns(df)
        
        # 计算支出数
        df = self._calculate_expenditure(df)
        
        # 提取单位编码并分组
        df = self._extract_unit_code_and_group(df)
        
        filtered_count = len(df)
        self.logger.info(f"会计核算数据处理完成: {original_count} -> {filtered_count} 条记录")
        
        return df
    
    def _clean_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """清洗数据"""
        return df[
            (df['账套'] != '借方合计') & 
            (df['账套'] != '贷方合计')
        ].copy()
    
    def _convert_numeric_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """转换数值列"""
        for col in ['借方累计', '贷方累计']:
            # 清理字符串并转换为数值
            df[col] = self.safe_numeric_conversion(
                self.clean_numeric_string(df[col])
            ).fillna(0)
        
        return df
    
    def _calculate_expenditure(self, df: pd.DataFrame) -> pd.DataFrame:
        """计算支出数"""
        df['会计核算_支出数'] = (df['借方累计'] - df['贷方累计']) / 10000
        return df
    
    def _extract_unit_code_and_group(self, df: pd.DataFrame) -> pd.DataFrame:
        """提取单位编码并分组"""
        df['单位编码'] = self.safe_numeric_conversion(
            df['账套'].str.slice(0, 6)
        )
        
        return df.groupby('单位编码', as_index=False)['会计核算_支出数'].sum()

class ExcelFormatter:
    """Excel格式化工具"""
    
    def __init__(self, logger: Logger):
        self.logger = logger
    
    def format_excel(self, file_path: Path, sheet_name: str, data_rows: int):
        """格式化Excel文件"""
        try:
            wb = load_workbook(file_path)
            ws = wb[sheet_name]
            
            self._apply_styles(ws)
            self._adjust_column_widths(ws)
            self._format_headers(ws)
            self._format_data_rows(ws, data_rows)
            self._add_features(ws, data_rows)
            self._add_summary_statistics(ws, data_rows)
            
            wb.save(file_path)
            self.logger.info(f"Excel格式化完成: {file_path}")
            
        except Exception as e:
            self.logger.error(f"Excel格式化失败: {str(e)}")
            raise
    
    def _apply_styles(self, ws):
        """应用样式"""
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # 为所有单元格添加边框
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
    
    def _adjust_column_widths(self, ws):
        """调整列宽"""
        max_column_width = ConfigManager.get_config('max_column_width')
        column_width_padding = ConfigManager.get_config('column_width_padding')
        
        column_widths = {}
        for row in ws.iter_rows():
            for i, cell in enumerate(row):
                width = len(str(cell.value or ''))
                column_widths[i] = max(column_widths.get(i, 0), width)
        
        for i, width in column_widths.items():
            column_letter = get_column_letter(i + 1)
            ws.column_dimensions[column_letter].width = min(
                width + column_width_padding, max_column_width
            )
    
    def _format_headers(self, ws):
        """格式化表头"""
        header_font = Font(name='微软雅黑', bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            cell.font = header_font
            cell.fill = header_fill
    
    def _format_data_rows(self, ws, data_rows: int):
        """格式化数据行"""
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        
        try:
            deviation_index = header.index('偏离度')
        except ValueError:
            self.logger.warning("未找到偏离度列，跳过偏离度格式化")
            return
        
        high_deviation_fill = PatternFill(
            start_color='FFFF00', end_color='FFFF00', fill_type='solid'
        )
        percentage_format = ConfigManager.get_config('percentage_format')
        high_deviation_threshold = ConfigManager.get_config('high_deviation_threshold')
        
        for row in ws.iter_rows(min_row=2, max_row=data_rows + 1):
            deviation_cell = row[deviation_index]
            
            if deviation_cell.value is not None:
                deviation_cell.number_format = percentage_format
                
                # 高偏离度行标黄
                if abs(deviation_cell.value) > high_deviation_threshold:
                    for cell in row:
                        cell.fill = high_deviation_fill
    
    def _add_features(self, ws, data_rows: int):
        """添加Excel功能"""
        # 冻结窗格
        ws.freeze_panes = ws['A2']
        
        # 自动筛选
        max_col = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f'A1:{max_col}{data_rows + 1}'
    
    def _add_summary_statistics(self, ws, data_rows: int):
        """添加汇总统计"""
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        
        try:
            deviation_index = header.index('偏离度')
            deviation_col = get_column_letter(deviation_index + 1)
        except ValueError:
            self.logger.warning("未找到偏离度列，跳过统计信息")
            return
        
        summary_row = data_rows + 3
        high_deviation_threshold = ConfigManager.get_config('high_deviation_threshold')
        
        # 添加统计信息
        statistics = [
            ('汇总统计', '', ''),
            ('总预算单位数', '', data_rows),
            (f'偏离度高于{high_deviation_threshold * 100:.0f}%的预算单位数', '', 
             f'=COUNTIF({deviation_col}2:{deviation_col}{data_rows + 1},">{high_deviation_threshold}")'
             f'+COUNTIF({deviation_col}2:{deviation_col}{data_rows + 1},"<{-high_deviation_threshold}")'),
            (f'生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', '', '')
        ]
        
        for i, (col_a, col_b, col_c) in enumerate(statistics):
            row_num = summary_row + i
            ws[f'A{row_num}'] = col_a
            ws[f'B{row_num}'] = col_b
            ws[f'C{row_num}'] = col_c
            
            if i == 0:  # 标题行加粗
                ws[f'A{row_num}'].font = Font(bold=True)

class AnalysisWorker(QThread):
    """分析工作线程"""
    progress = pyqtSignal(int)
    finished = pyqtSignal(str)
    error = pyqtSignal(str)
    
    def __init__(self, analyzer):
        super().__init__()
        self.analyzer = analyzer
    
    def run(self):
        try:
            self.progress.emit(10)
            self.analyzer.read_data()
            self.progress.emit(40)
            
            self.analyzer.process_data()
            self.progress.emit(70)
            
            output_path = self.analyzer.save_results()
            self.progress.emit(100)
            
            self.finished.emit(output_path)
        except Exception as e:
            self.error.emit(str(e))

class AccountingAnalyzer:
    """会计核算数据与预算执行数据对比分析工具"""
    
    def __init__(self, yszx_path: str = '', 
                 kjhs_path: str = '', 
                 output_path: str = ''):
        """初始化分析器"""
        self.yszx_path = Path(yszx_path) if yszx_path else Path('./导出数据')
        self.kjhs_path = Path(kjhs_path) if kjhs_path else Path('./会计核算.xls')
        
        # 如果未指定输出路径，生成带时间戳的文件名
        if output_path:
            self.output_path = Path(output_path)
        else:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            self.output_path = Path(f'./会计核算财政资金偏离度_{timestamp}.xlsx')
        
        # 暂时不初始化logger，等待UI创建后再初始化
        self.logger = None
        self.budget_processor = None
        self.accounting_processor = None
        self.excel_formatter = None
        
        self.yszx_df = None
        self.kjhs_df = None
        self.result_df = None
        
        # 记录生成时间，用于后续引用
        self.generation_timestamp = datetime.now()
    
    def set_logger(self, logger: Logger):
        """设置日志记录器并初始化处理器"""
        self.logger = logger
        self.budget_processor = BudgetExecutionProcessor(self.logger)
        self.accounting_processor = AccountingProcessor(self.logger)
        self.excel_formatter = ExcelFormatter(self.logger)
    
    def read_data(self) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """读取数据文件"""
        self.logger.info("开始读取数据...")
        
        # 读取预算执行数据
        yszx_df = self._read_file(self.yszx_path, ['xlsx', 'csv'])
        
        # 读取会计核算数据，跳过前4行
        kjhs_df = self._read_file(self.kjhs_path, ['xls', 'xlsx'], skiprows=4)
        
        self.yszx_df = yszx_df
        self.kjhs_df = kjhs_df
        
        return yszx_df, kjhs_df
    
    def _read_file(self, file_path: Path, valid_extensions: list, skiprows: int = 0) -> pd.DataFrame:
        """通用文件读取方法"""
        for ext in valid_extensions:
            full_path = file_path.with_suffix(f'.{ext}')
            if full_path.exists():
                if ext == 'xlsx' or ext == 'xls':
                    df = pd.read_excel(full_path, skiprows=skiprows)
                elif ext == 'csv':
                    df = pd.read_csv(full_path)
                self.logger.info(f"成功读取数据: {full_path}")
                return df
        
        raise FileNotFoundError(f"未找到有效数据文件: {file_path}")
    
    def process_data(self) -> pd.DataFrame:
        """处理数据"""
        if self.yszx_df is None or self.kjhs_df is None:
            raise ValueError("请先读取数据")
        
        # 处理预算执行数据
        processed_yszx = self.budget_processor.process(self.yszx_df)
        
        # 处理会计核算数据
        processed_kjhs = self.accounting_processor.process(self.kjhs_df)
        
        # 合并数据并计算偏离度
        result_df = self._merge_and_calculate_deviation(processed_yszx, processed_kjhs)
        
        self.result_df = result_df
        return result_df
    
    def _merge_and_calculate_deviation(self, yszx_df: pd.DataFrame, 
                                     kjhs_df: pd.DataFrame) -> pd.DataFrame:
        """合并数据并计算偏离度"""
        self.logger.info("开始合并数据并计算偏离度...")
        
        # 合并数据
        merged_df = pd.merge(yszx_df, kjhs_df, on='单位编码', how='left')
        merged_df['会计核算_支出数'] = merged_df['会计核算_支出数'].fillna(0)
        
        # 计算差额和偏离度
        merged_df['差额'] = merged_df['会计核算_支出数'] - merged_df['预算执行_支出数']
        merged_df['差额'] = merged_df['差额'].round(
            ConfigManager.get_config('decimal_places')
        )
        
        # 安全计算偏离度
        merged_df['偏离度'] = merged_df.apply(self._calculate_deviation_safe, axis=1)
        merged_df['备注'] = ''
        
        # 重新排列列
        result_df = merged_df[[
            '序号', '单位编码', '预算单位', '预算执行_支出数', 
            '会计核算_支出数', '差额', '偏离度', '备注'
        ]]
        
        self.logger.info(f"数据合并完成，共 {len(result_df)} 条记录")
        return result_df
    
    @staticmethod
    def _calculate_deviation_safe(row) -> Optional[float]:
        """安全计算偏离度"""
        if pd.notna(row['预算执行_支出数']) and row['预算执行_支出数'] != 0:
            return row['差额'] / row['预算执行_支出数']
        return None
    
    def save_results(self):
        """保存结果"""
        if self.result_df is None:
            raise ValueError("请先处理数据")
        
        self.logger.info(f"开始保存结果到: {self.output_path}")
        
        # 确保输出目录存在
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # 保存到Excel
        self.result_df.to_excel(self.output_path, sheet_name='偏离度', index=False)
        
        # 格式化Excel
        self.excel_formatter.format_excel(
            self.output_path, '偏离度', len(self.result_df)
        )
        
        self.logger.info(f"分析结果已保存到: {self.output_path}")
        return str(self.output_path)
    
    def run_analysis(self):
        """执行完整分析流程"""
        #self.logger.info("#" * 35)
        self.logger.info("开始会计核算与预算执行数据对比分析")
        self.logger.info(f"输出文件: {self.output_path}")
        # self.logger.info("#" * 35)
        
        self.read_data()
        # self.logger.info("=" * 35)

        self.process_data()
        # self.logger.info("=" * 35)

        output_path = self.save_results()
        # self.logger.info("*" * 35)

        self.logger.info("分析完成")
        # self.logger.info("*" * 35)
        
        return output_path

class MainWindow(QMainWindow):
    """主窗口类"""
    log_signal = pyqtSignal(str)

    def __init__(self, logger: logging.Logger = None):
        super().__init__()
        self.setWindowTitle("计算会计核算偏离度工具")
        self.setGeometry(100, 100, 800, 600)
        self.setMinimumSize(600, 500)

        # 初始化分析器
        self.analyzer = AccountingAnalyzer()

        # 使用外部注入的 logger（优先）或在本地创建
        self.logger = logger
        try:
            if self.logger is None:
                # 没有传入 logger 时，按原逻辑创建一个并将 qt 信号附加到该 logger
                self.logger = get_logger('AccountingAnalyzer', qt_signal=self.log_signal)
            else:
                # 如果传入了 logger，则把本窗口的 signal 附加到该 logger
                add_qt_signal(self.logger, self.log_signal)
        except Exception:
            # 在任何错误情况下回退为空 logger
            self.logger = None

        self.analyzer.set_logger(self.logger)
        self.log_signal.connect(self.update_log)
        
        # 创建UI
        self.init_ui()
        
        # 发送应用程序启动消息
        self.logger.info("应用程序已启动")
        
        # 初始化工作线程
        self.worker = None
    
    def init_ui(self):
        """初始化用户界面"""
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        main_layout = QVBoxLayout(central_widget)
        
        # 使用全局样式表 (common/style.qss) 加载于应用入口处，避免模块内重复样式定义
        # 如需模块特定样式，可使用 setObjectName / setProperty 并在 common/style.qss 中添加选择器
        central_widget.setObjectName('accounting_central')
        
        # 创建配置区域
        self.create_config_section(main_layout)
        
        # 创建文件选择区域
        self.create_file_selection_section(main_layout)
        
        # 创建进度条
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        main_layout.addWidget(self.progress_bar)
        
        # 创建按钮区域
        self.create_buttons_section(main_layout)
        
        # 创建日志显示区域
        self.create_log_section(main_layout)
        
        # 设置默认输出路径
        self.update_default_output_path()
    
    def create_config_section(self, layout):
        """创建配置区域"""
        config_group = QGroupBox("分析配置")
        config_layout = QGridLayout(config_group)
        
        # 高偏离度阈值设置
        threshold_label = QLabel("偏离度阈值 (%):")
        self.deviation_threshold_edit = QLineEdit(str(ConfigManager.get_config('high_deviation_threshold') * 100))
        self.deviation_threshold_edit.setFixedWidth(100)
        
        # 保留小数位数设置
        decimal_label = QLabel("差额列保留小数位数:")
        self.decimal_places_edit = QLineEdit(str(ConfigManager.get_config('decimal_places')))
        self.decimal_places_edit.setFixedWidth(50)
        
        # 应用配置按钮
        apply_config_btn = QPushButton("1.应用配置")
        apply_config_btn.clicked.connect(self.apply_config)
        apply_config_btn.setFixedWidth(100)
        
        config_layout.addWidget(threshold_label, 0, 0)
        config_layout.addWidget(self.deviation_threshold_edit, 0, 1)
        config_layout.addWidget(decimal_label, 0, 2)
        config_layout.addWidget(self.decimal_places_edit, 0, 3)
        config_layout.addWidget(apply_config_btn, 0, 4)
        
        # 配置说明
        config_note = QLabel("*修改配置将影响计算结果")
        config_note.setFont(QFont("Microsoft YaHei", 9))
        config_layout.addWidget(config_note, 1, 0, 1, 5)
        
        layout.addWidget(config_group)
    
    def create_file_selection_section(self, layout):
        """创建文件选择区域"""
        file_group = QGroupBox("文件路径")
        file_layout = QGridLayout(file_group)
        
        # 预算执行数据路径
        yszx_label = QLabel("2.预算执行数据:")
        self.yszx_path_edit = QLineEdit()
        yszx_browse_btn = QPushButton("浏览...")
        yszx_browse_btn.clicked.connect(self.browse_yszx_file)
        
        # 会计核算数据路径
        kjhs_label = QLabel("3.会计核算数据:")
        self.kjhs_path_edit = QLineEdit()
        kjhs_browse_btn = QPushButton("浏览...")
        kjhs_browse_btn.clicked.connect(self.browse_kjhs_file)
        
        # 输出文件路径
        output_label = QLabel("输出文件路径:")
        self.output_path_edit = QLineEdit()
        output_browse_btn = QPushButton("浏览...")
        output_browse_btn.clicked.connect(self.browse_output_file)
        
        # 添加到布局
        file_layout.addWidget(yszx_label, 0, 0)
        file_layout.addWidget(self.yszx_path_edit, 0, 1)
        file_layout.addWidget(yszx_browse_btn, 0, 2)
        
        file_layout.addWidget(kjhs_label, 1, 0)
        file_layout.addWidget(self.kjhs_path_edit, 1, 1)
        file_layout.addWidget(kjhs_browse_btn, 1, 2)
        
        file_layout.addWidget(output_label, 2, 0)
        file_layout.addWidget(self.output_path_edit, 2, 1)
        file_layout.addWidget(output_browse_btn, 2, 2)
        
        layout.addWidget(file_group)
    
    def create_buttons_section(self, layout):
        """创建按钮区域"""
        buttons_layout = QHBoxLayout()
        
        # 开始分析按钮
        self.analyze_btn = QPushButton("4.分析并导出")
        self.analyze_btn.clicked.connect(self.start_analysis)
        
        # 退出按钮
        exit_btn = QPushButton("退出")
        exit_btn.clicked.connect(self.close)
        
        buttons_layout.addStretch()
        buttons_layout.addWidget(self.analyze_btn)
        buttons_layout.addWidget(exit_btn)
        
        layout.addLayout(buttons_layout)
    
    def create_log_section(self, layout):
        """创建日志显示区域"""
        log_group = QGroupBox("日志信息")
        log_layout = QVBoxLayout(log_group)
        
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        
        log_layout.addWidget(self.log_text)
        layout.addWidget(log_group)
    
    def update_default_output_path(self):
        """更新默认输出路径"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_output = f"./会计核算财政资金偏离度_{timestamp}.xlsx"
        self.output_path_edit.setText(default_output)
    
    def browse_yszx_file(self):
        """浏览预算执行数据文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择预算执行数据文件",
            "", "Excel files (*.xlsx *.xls);;CSV files (*.csv);;All files (*.*)"
        )
        if file_path:
            self.yszx_path_edit.setText(file_path)
            self.analyzer.yszx_path = Path(file_path)
    
    def browse_kjhs_file(self):
        """浏览会计核算数据文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择会计核算数据文件",
            "", "Excel files (*.xlsx *.xls);;All files (*.*)"
        )
        if file_path:
            self.kjhs_path_edit.setText(file_path)
            self.analyzer.kjhs_path = Path(file_path)
    
    def browse_output_file(self):
        """浏览输出文件路径"""
        file_path, _ = QFileDialog.getSaveFileName(
            self, "保存分析结果",
            self.output_path_edit.text(), "Excel files (*.xlsx)"
        )
        if file_path:
            self.output_path_edit.setText(file_path)
            self.analyzer.output_path = Path(file_path)
    
    def apply_config(self):
        """应用配置更改"""
        try:
            # 转换高偏离度阈值
            threshold_text = self.deviation_threshold_edit.text()
            threshold = float(threshold_text) / 100
            if threshold <= 0:
                raise ValueError("阈值必须大于0")
            ConfigManager.set_config('high_deviation_threshold', threshold)
            
            # 转换小数位数
            decimals = int(self.decimal_places_edit.text())
            if decimals < 0 or decimals > 10:
                raise ValueError("小数位数必须在0-10之间")
            ConfigManager.set_config('decimal_places', decimals)
            
            # 使用原始输入文本进行日志记录，避免浮点数精度问题
            self.logger.info(f"配置已更新 - 高偏离度阈值: {threshold_text}%, 小数位数: {decimals}")
        except ValueError as e:
            self.logger.error(f"配置错误: {str(e)}")
            QMessageBox.critical(self, "配置错误", str(e))
    
    def start_analysis(self):
        """开始分析过程"""
        # 禁用分析按钮防止重复点击
        self.analyze_btn.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.logger.info("准备开始分析...")
        
        # 更新分析器路径
        self.analyzer.yszx_path = Path(self.yszx_path_edit.text())
        self.analyzer.kjhs_path = Path(self.kjhs_path_edit.text())
        self.analyzer.output_path = Path(self.output_path_edit.text())
        
        # 创建并启动工作线程
        self.worker = AnalysisWorker(self.analyzer)
        self.worker.progress.connect(self.update_progress)
        self.worker.finished.connect(self.analysis_completed)
        self.worker.error.connect(self.analysis_failed)
        self.worker.start()
    
    def update_progress(self, value):
        """更新进度条"""
        self.progress_bar.setValue(value)
    
    def analysis_completed(self, output_path):
        """分析完成回调"""
        self.analyze_btn.setEnabled(True)
        self.progress_bar.setVisible(False)
        QMessageBox.information(self, "成功", f"分析已完成！结果已保存至:\n{output_path}")
    
    def analysis_failed(self, error_msg):
        """分析失败回调"""
        self.analyze_btn.setEnabled(True)
        self.progress_bar.setVisible(False)
        QMessageBox.critical(self, "分析失败", f"分析过程中发生错误:\n{error_msg}")
    
    def update_log(self, message):
        """更新日志显示"""
        self.log_text.append(message)
        # 自动滚动到底部
        self.log_text.verticalScrollBar().setValue(self.log_text.verticalScrollBar().maximum())

def main():
    """主函数"""
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())

if __name__ == "__main__":
    main()